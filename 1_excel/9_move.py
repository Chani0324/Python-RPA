from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

# ws.move_range("B1:C11", rows=0, cols=1) # 이동시킬 범위, 얼마만큼 row, col을 이동시킬건지
# ws["B1"].value = "국어"

# 번호 영어 수학
ws.move_range("C1:C11", rows=5, cols=-1)





wb.save("sample_korean.xlsx")

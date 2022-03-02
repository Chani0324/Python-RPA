from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 영어 column만 가지고 오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)
    
col_range = ws["B:C"] # 영어, 수학 column 함께 가지고 오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1] # 1번째 row만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#         print()



# row_range = ws[2:ws.max_row] # 2번쨰 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ") 
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="") # A, B, ...
#         print(xy[1], end=" ") # 1, 2, ...
#     print()


# 전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[1].value)

# 전체 coulmns
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)

# for column in ws.iter_cols(): # 전체 column
#     print(column[0].value)

# for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3): 
    # print(row[0].value, row[1].value)
#     print(row)

# for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
#     print(col)


wb.save("sample.xlsx")
